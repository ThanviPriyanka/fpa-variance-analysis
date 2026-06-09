import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.worksheet.properties import PageSetupProperties

SEGS = ["United Kingdom", "EIRE", "Netherlands", "Germany", "France", "Other"]
N = len(SEGS)

INK   = "29333B"; TEAL = "0E7C7B"; FAV = "2E7D32"; ADV = "C0392B"
TINT  = "E8EEEE"; GREY = "9AA5A5"; ARIAL = "Arial"
GBP  = '£#,##0;(£#,##0);-'; GBP2 = '£#,##0.00;(£#,##0.00);-'
NUM  = '#,##0;(#,##0);-'; PCT = '0.0%'

BLUE  = Font(name=ARIAL, color="0000FF")
BLACK = Font(name=ARIAL, color="000000")
LINKF = Font(name=ARIAL, color=TEAL)
TITLE = Font(name=ARIAL, size=15, bold=True, color=INK)
H2    = Font(name=ARIAL, size=11, bold=True, color=TEAL)
WHITEB= Font(name=ARIAL, bold=True, color="FFFFFF")
BOLD  = Font(name=ARIAL, bold=True, color=INK)
INKFILL = PatternFill("solid", fgColor=INK)
TINTFILL= PatternFill("solid", fgColor=TINT)
YEL   = PatternFill("solid", fgColor="FFF2CC")
thin  = Side(style="thin", color="C9D2D2")
BORD  = Border(top=thin, bottom=thin, left=thin, right=thin)
topb  = Border(top=Side(style="thin", color=INK))

def hdr(ws, row, cols, start=1):
    for i, c in enumerate(cols):
        cell = ws.cell(row=row, column=start + i, value=c)
        cell.font = WHITEB; cell.fill = INKFILL
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BORD

b_t, b_h, b_f = 3, 4, 5
b_l = b_f + N - 1; b_tot = b_l + 1
bu_t = b_tot + 2; bu_h = bu_t + 1; bu_f = bu_h + 1; bu_l = bu_f + N - 1; bu_tot = bu_l + 1
a_t = bu_tot + 2; a_h = a_t + 1; a_f = a_h + 1; a_l = a_f + N - 1; a_tot = a_l + 1
br_t = a_tot + 2; br_f = br_t + 1

wb = Workbook()

fact = pd.read_csv("data/processed/monthly_fact.csv")
wsd = wb.active; wsd.title = "Data"
wsd["A1"] = "MONTHLY FACT TABLE  —  output of sql/build_fact_table_by_market.sql (DuckDB)"; wsd["A1"].font = TITLE
wsd["A2"] = "Imported from the data pipeline (real Online Retail II). Re-run the SQL to refresh; do not edit by hand."
wsd["A2"].font = Font(name=ARIAL, italic=True, color="808080")
hdr(wsd, 4, ["year", "month", "category", "units", "revenue", "avg_price"])
for r, row in enumerate(fact.itertuples(index=False), start=5):
    wsd.cell(r, 1, int(row.year)).font = BLACK
    wsd.cell(r, 2, int(row.month)).font = BLACK
    wsd.cell(r, 3, row.category).font = BLACK
    wsd.cell(r, 4, int(row.units)).font = BLACK; wsd.cell(r, 4).number_format = NUM
    wsd.cell(r, 5, float(row.revenue)).font = BLACK; wsd.cell(r, 5).number_format = GBP
    wsd.cell(r, 6, float(row.avg_price)).font = BLACK; wsd.cell(r, 6).number_format = GBP2
for col, w in zip("ABCDEF", [8, 8, 16, 12, 14, 12]):
    wsd.column_dimensions[col].width = w
wsd.freeze_panes = "A5"

wsa = wb.create_sheet("Assumptions")
wsa["A1"] = "FP&A Variance Model  —  Assumptions (the plan, set in advance)"; wsa["A1"].font = TITLE
wsa["A3"] = "Base year (actuals the plan is built from)"; wsa["A3"].font = BOLD
wsa["B3"] = 2010; wsa["B3"].font = BLUE; wsa["B3"].fill = YEL; wsa["B3"].number_format = "0"
wsa["A4"] = "Budget year (plan vs actual compared)"; wsa["A4"].font = BOLD
wsa["B4"] = 2011; wsa["B4"].font = BLUE; wsa["B4"].fill = YEL; wsa["B4"].number_format = "0"
wsa["A6"] = "Blue / shaded cells are inputs. Change them and the whole model + bridge update."
wsa["A6"].font = Font(name=ARIAL, italic=True, color="808080")
wsa["A7"] = "Comparison is like-for-like Jan–Nov (December 2011 is truncated in the source, so it is excluded)."
wsa["A7"].font = Font(name=ARIAL, italic=True, color="808080")
hdr(wsa, 8, ["Market segment", "Volume Growth %", "Price Change %"])
for i, seg in enumerate(SEGS):
    r = 9 + i
    wsa.cell(r, 1, seg).font = BLACK; wsa.cell(r, 1).border = BORD
    for col in (2, 3):
        c = wsa.cell(r, col, 0.10 if col == 2 else 0.03)
        c.font = BLUE; c.fill = YEL; c.number_format = PCT; c.border = BORD
for col, w in zip("ABC", [40, 16, 16]):
    wsa.column_dimensions[col].width = w

wsm = wb.create_sheet("Model")
MC = 'Data!$B:$B,"<=11"'
wsm["A1"] = "Calculation engine  —  Base → Budget → Actual → Price/Volume/Mix bridge"; wsm["A1"].font = TITLE
wsm["A2"] = "All figures are Jan–Nov (like-for-like). Segment = market."; wsm["A2"].font = Font(name=ARIAL, italic=True, color="808080")

wsm.cell(b_t, 1, "1.  Base year actuals  (Jan–Nov, = Assumptions base year)").font = H2
hdr(wsm, b_h, ["Market", "Units", "Revenue", "Avg Price"])
for i, seg in enumerate(SEGS):
    r = b_f + i
    wsm.cell(r, 1, seg).font = BLACK; wsm.cell(r, 1).border = BORD
    wsm.cell(r, 2, f"=SUMIFS(Data!$D:$D,Data!$A:$A,Assumptions!$B$3,Data!$C:$C,$A{r},{MC})").font = LINKF
    wsm.cell(r, 3, f"=SUMIFS(Data!$E:$E,Data!$A:$A,Assumptions!$B$3,Data!$C:$C,$A{r},{MC})").font = LINKF
    wsm.cell(r, 4, f"=C{r}/B{r}").font = BLACK
    wsm.cell(r, 2).number_format = NUM; wsm.cell(r, 3).number_format = GBP; wsm.cell(r, 4).number_format = GBP2
wsm.cell(b_tot, 1, "Total").font = BOLD
wsm.cell(b_tot, 2, f"=SUM(B{b_f}:B{b_l})"); wsm.cell(b_tot, 3, f"=SUM(C{b_f}:C{b_l})"); wsm.cell(b_tot, 4, f"=C{b_tot}/B{b_tot}")
for col in "BCD": wsm[f"{col}{b_tot}"].font = BOLD; wsm[f"{col}{b_tot}"].border = topb
wsm[f"B{b_tot}"].number_format = NUM; wsm[f"C{b_tot}"].number_format = GBP; wsm[f"D{b_tot}"].number_format = GBP2

wsm.cell(bu_t, 1, "2.  Budget  (base × plan assumptions)").font = H2
hdr(wsm, bu_h, ["Market", "Vol Gr %", "Price Ch %", "Bud Units", "Bud Price", "Bud Revenue"])
for i, seg in enumerate(SEGS):
    r = bu_f + i; base = b_f + i; asum = 9 + i
    wsm.cell(r, 1, f"=A{base}").font = BLACK; wsm.cell(r, 1).border = BORD
    wsm.cell(r, 2, f"=Assumptions!B{asum}").font = LINKF; wsm.cell(r, 2).number_format = PCT
    wsm.cell(r, 3, f"=Assumptions!C{asum}").font = LINKF; wsm.cell(r, 3).number_format = PCT
    wsm.cell(r, 4, f"=B{base}*(1+B{r})").font = BLACK; wsm.cell(r, 4).number_format = NUM
    wsm.cell(r, 5, f"=D{base}*(1+C{r})").font = BLACK; wsm.cell(r, 5).number_format = GBP2
    wsm.cell(r, 6, f"=D{r}*E{r}").font = BLACK; wsm.cell(r, 6).number_format = GBP
wsm.cell(bu_tot, 1, "Total").font = BOLD
wsm.cell(bu_tot, 4, f"=SUM(D{bu_f}:D{bu_l})"); wsm.cell(bu_tot, 6, f"=SUM(F{bu_f}:F{bu_l})")
for col in "DF": wsm[f"{col}{bu_tot}"].font = BOLD; wsm[f"{col}{bu_tot}"].border = topb
wsm[f"D{bu_tot}"].number_format = NUM; wsm[f"F{bu_tot}"].number_format = GBP

wsm.cell(a_t, 1, "3.  Actual  (Jan–Nov, = Assumptions budget year)  +  per-segment effects").font = H2
hdr(wsm, a_h, ["Market", "Act Units", "Act Revenue", "Act Price", "Mix effect", "Price effect"])
for i, seg in enumerate(SEGS):
    r = a_f + i; base = b_f + i; bud = bu_f + i
    wsm.cell(r, 1, f"=A{base}").font = BLACK; wsm.cell(r, 1).border = BORD
    wsm.cell(r, 2, f"=SUMIFS(Data!$D:$D,Data!$A:$A,Assumptions!$B$4,Data!$C:$C,$A{r},{MC})").font = LINKF
    wsm.cell(r, 3, f"=SUMIFS(Data!$E:$E,Data!$A:$A,Assumptions!$B$4,Data!$C:$C,$A{r},{MC})").font = LINKF
    wsm.cell(r, 4, f"=C{r}/B{r}").font = BLACK
    wsm.cell(r, 5, f"=(B{r}-D{bud}*($B${a_tot}/$D${bu_tot}))*E{bud}").font = BLACK
    wsm.cell(r, 6, f"=(D{r}-E{bud})*B{r}").font = BLACK
    wsm.cell(r, 2).number_format = NUM; wsm.cell(r, 3).number_format = GBP
    wsm.cell(r, 4).number_format = GBP2; wsm.cell(r, 5).number_format = GBP; wsm.cell(r, 6).number_format = GBP
wsm.cell(a_tot, 1, "Total").font = BOLD
wsm.cell(a_tot, 2, f"=SUM(B{a_f}:B{a_l})"); wsm.cell(a_tot, 3, f"=SUM(C{a_f}:C{a_l})")
wsm.cell(a_tot, 5, f"=SUM(E{a_f}:E{a_l})"); wsm.cell(a_tot, 6, f"=SUM(F{a_f}:F{a_l})")
for col in "BCEF": wsm[f"{col}{a_tot}"].font = BOLD; wsm[f"{col}{a_tot}"].border = topb
wsm[f"B{a_tot}"].number_format = NUM
for col in "CEF": wsm[f"{col}{a_tot}"].number_format = GBP

wsm.cell(br_t, 1, "4.  Revenue bridge:  Budget → Volume → Mix → Price → Actual").font = H2
bridge = [("Budget revenue", f"=F{bu_tot}", True), ("Volume effect", f"=(B{a_tot}-D{bu_tot})*(F{bu_tot}/D{bu_tot})", False),
          ("Mix effect", f"=E{a_tot}", False), ("Price effect", f"=F{a_tot}", False),
          ("Actual revenue", f"=C{a_tot}", True), ("Total variance", f"=C{a_tot}-F{bu_tot}", True),
          ("Bridge check (≈0)", f"=(C{a_tot}-F{bu_tot})-(B{br_f+1}+B{br_f+2}+B{br_f+3})", False)]
for i, (lbl, f, bold) in enumerate(bridge):
    r = br_f + i
    wsm.cell(r, 1, lbl).font = BOLD if bold else BLACK
    c = wsm.cell(r, 2, f); c.font = BOLD if bold else BLACK; c.number_format = GBP
wsm.cell(br_f + 6, 2).number_format = '£#,##0.00'
for col, w in zip("ABCDEF", [30, 12, 14, 12, 12, 14]):
    wsm.column_dimensions[col].width = w

wsb = wb.create_sheet("Dashboard")
wsb.sheet_view.showGridLines = False
wsb["A1"] = "FY2011 Revenue Variance  —  Plan vs Actual  (by market)"; wsb["A1"].font = Font(name=ARIAL, size=16, bold=True, color=INK)
wsb["A2"] = ("Finding: 2011 revenue grew ~2% year-over-year but landed ~10% below a +10% growth plan. The gap is almost "
             "entirely a volume shortfall (the growth assumption was too aggressive); better-than-planned pricing cushioned it. "
             "Like-for-like Jan–Nov — December 2011 is truncated in the source and excluded.")
wsb["A2"].font = Font(name=ARIAL, size=10, italic=True, color="566066")
wsb.merge_cells("A2:H2"); wsb["A2"].alignment = Alignment(wrap_text=True, vertical="top")
wsb.row_dimensions[2].height = 42

kpis = [("Budget revenue", f"=Model!F{bu_tot}", GBP, INK), ("Actual revenue", f"=Model!C{a_tot}", GBP, INK),
        ("Total variance", f"=Model!C{a_tot}-Model!F{bu_tot}", GBP, ADV), ("Variance %", f"=(Model!C{a_tot}-Model!F{bu_tot})/Model!F{bu_tot}", PCT, ADV)]
for i, (lbl, f, fmt, valcolor) in enumerate(kpis):
    col = 1 + i * 2
    lc = wsb.cell(4, col, lbl); lc.font = WHITEB; lc.fill = INKFILL
    lc.alignment = Alignment(horizontal="center"); wsb.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    vc = wsb.cell(5, col, f); vc.font = Font(name=ARIAL, size=14, bold=True, color=valcolor)
    vc.number_format = fmt; vc.fill = TINTFILL; vc.alignment = Alignment(horizontal="center")
    wsb.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1); wsb.cell(5, col + 1).fill = TINTFILL
wsb.row_dimensions[5].height = 26

wsb["J3"] = "Helper tables (drive the charts)"; wsb["J3"].font = Font(name=ARIAL, size=9, color="808080")
hdr(wsb, 4, ["Step", "Value", "CumEnd", "Base", "Decrease", "Increase", "Total"], start=10)
steps = [("Budget", f"=Model!B{br_f}"), ("Volume", f"=Model!B{br_f+1}"), ("Mix", f"=Model!B{br_f+2}"),
         ("Price", f"=Model!B{br_f+3}"), ("Actual", f"=Model!B{br_f+4}")]
for i, (lbl, vf) in enumerate(steps):
    r = 5 + i
    wsb.cell(r, 10, lbl).font = BLACK; wsb.cell(r, 11, vf).font = BLACK
    if lbl in ("Budget", "Actual"):
        wsb.cell(r, 12, f"=K{r}"); wsb.cell(r, 13, 0); wsb.cell(r, 14, 0); wsb.cell(r, 15, 0); wsb.cell(r, 16, f"=K{r}")
    else:
        wsb.cell(r, 12, f"=L{r-1}+K{r}"); wsb.cell(r, 13, f"=L{r}-MAX(K{r},0)")
        wsb.cell(r, 14, f"=MAX(-K{r},0)"); wsb.cell(r, 15, f"=MAX(K{r},0)"); wsb.cell(r, 16, 0)
    for col in range(11, 17):
        wsb.cell(r, col).number_format = GBP; wsb.cell(r, col).font = BLACK

wf = BarChart(); wf.type = "col"; wf.grouping = "stacked"; wf.overlap = 100
wf.title = "Revenue Bridge: Budget → Actual (£, Jan–Nov)"; wf.height = 8.5; wf.width = 17
for col in (13, 14, 15, 16):
    wf.add_data(Reference(wsb, min_col=col, min_row=4, max_row=9), titles_from_data=True)
wf.set_categories(Reference(wsb, min_col=10, min_row=5, max_row=9))
wf.series[0].graphicalProperties.noFill = True
wf.series[0].graphicalProperties.line.noFill = True
wf.series[1].graphicalProperties.solidFill = ADV
wf.series[2].graphicalProperties.solidFill = FAV
wf.series[3].graphicalProperties.solidFill = INK
wf.legend = None; wf.y_axis.numFmt = GBP; wf.y_axis.majorGridlines = None
wsb.add_chart(wf, "A7")

hdr(wsb, 12, ["Month", "Budget", "Actual"], start=10)
for m in range(1, 12):
    r = 12 + m
    wsb.cell(r, 10, m).font = BLACK
    wsb.cell(r, 11, f"=SUMIFS(Data!$E:$E,Data!$A:$A,Assumptions!$B$3,Data!$B:$B,J{r})*(Model!$F${bu_tot}/Model!$C${b_tot})").font = BLACK
    wsb.cell(r, 12, f"=SUMIFS(Data!$E:$E,Data!$A:$A,Assumptions!$B$4,Data!$B:$B,J{r})").font = BLACK
    wsb.cell(r, 11).number_format = GBP; wsb.cell(r, 12).number_format = GBP

ln = LineChart(); ln.title = "Monthly Revenue: Budget vs Actual (£, Jan–Nov)"; ln.height = 8.5; ln.width = 17
ln.add_data(Reference(wsb, min_col=11, min_row=12, max_col=12, max_row=23), titles_from_data=True)
ln.set_categories(Reference(wsb, min_col=10, min_row=13, max_row=23))
ln.series[0].graphicalProperties.line.solidFill = GREY
ln.series[1].graphicalProperties.line.solidFill = TEAL
ln.series[1].graphicalProperties.line.width = 30000
ln.y_axis.numFmt = GBP; ln.x_axis.title = "Month"
wsb.add_chart(ln, "A26")

for col, w in zip("ABCDEFGH", [16, 13, 13, 13, 13, 13, 13, 13]):
    wsb.column_dimensions[col].width = w
for col in "JKLMNOP":
    wsb.column_dimensions[col].width = 11
wsb.page_setup.orientation = "landscape"
wsb.page_setup.fitToWidth = 1; wsb.page_setup.fitToHeight = 1
wsb.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
wsb.print_area = "A1:H44"

wsn = wb.create_sheet("Notes")
wsn["A1"] = "Methodology & Notes"; wsn["A1"].font = TITLE
notes = ["", "DATA SOURCE",
 "• UCI 'Online Retail II' — real UK online-retail transactions, Dec 2009 – Dec 2011 (~1.07M rows).",
 "• Pipeline: sql/build_fact_table_by_market.sql (DuckDB).", "",
 "CLEANING  (Stage 2 profiling on 1,067,371 raw rows)",
 "• Dropped cancellations ('C' invoices), non-positive quantities, and non-positive prices (~2.4%).",
 "• Kept 1,041,670 rows (97.6%). Missing Customer IDs (243,007) retained — revenue does not need them.", "",
 "SEGMENTATION  (v1 'by market' cut)",
 "• Segment = market: United Kingdom + EIRE, Netherlands, Germany, France, and 'Other' (all remaining countries).",
 "• The UK is ~85% of revenue, so cross-market mix is a minor variance driver by construction.",
 "• A v2 'by product category' cut (derived from descriptions) would show a richer mix story.", "",
 "COMPARISON PERIOD",
 "• Like-for-like Jan–Nov 2010 vs Jan–Nov 2011. December 2011 is truncated in the source (data ends",
 "  mid-December), so including it would invent a false ~£0.6M shortfall. December is excluded from both years.", "",
 "FORECAST / BUDGET",
 "• Driver-based: Budget = base-year actuals × (1 + Volume Growth %) × (1 + Price Change %), per segment.",
 "• Plan set in advance on the Assumptions tab (default +10% volume, +3% price). All inputs editable.", "",
 "PRICE / VOLUME / MIX BRIDGE  (decomposes Actual − Budget revenue variance)",
 "• Volume = (Total Actual units − Total Budget units) × Budget weighted-avg price",
 "• Mix    = Σ [ Actual units_i − Budget units_i × (Total Actual / Total Budget) ] × Budget price_i",
 "• Price  = Σ ( Actual price_i − Budget price_i ) × Actual units_i",
 "• Volume + Mix + Price = total variance (see 'Bridge check' on the Model tab — it equals ~£0).", "",
 "HONESTY NOTE",
 "• The budget is a constructed plan for this exercise; the finding is conditional on the plan assumptions.",
 "• Change the Assumptions and the variance — and the story — change accordingly."]
for i, t in enumerate(notes):
    c = wsn.cell(2 + i, 1, t)
    c.font = H2 if (t.isupper() and t) else Font(name=ARIAL, size=10, color="404040")
wsn.column_dimensions["A"].width = 105

wb._sheets.sort(key=lambda s: ["Dashboard", "Model", "Assumptions", "Data", "Notes"].index(s.title))
wb.active = 0
wb.save("model/FPA_Variance_Model.xlsx")
print("saved")
